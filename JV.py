import openpyxl as pyxl
import pyautogui
import win32com.client as win32
import os
import csv
import ctypes
from datetime import datetime
from tkinter import *
from tkinter import filedialog

#this function is used to update the error text file
def writeErrorFile(file, booleanText):
	file.seek(0)
	file.truncate()
	file.write(booleanText)
	#reset read position to first line 
	file.seek(0)
#this function updates the log file
def updateLogFile(msgArr):
	with open('log.txt', 'a') as file:
		today = datetime.today()
		current_day = today.strftime("%m/%d/%y")
		now = datetime.now()
		current_time = now.strftime("%H:%M:%S")
		file.write(current_day + " " + current_time + "\n")
		for msg in msgArr:
			file.write(msg + "\n")
		file.write("\n")
#this function checks if a file exists in a certain directory, with an array of certain extensions
#it returns a dictionary of information about a file (if its .xls or .xlsx and if it exists)
def checkFileExist(fileName, directory, extensions):
	fileExistInfo = {"exist": False, "xlsx": False}
	#check if the file exists given by the user
	for dirName, subdirList, fileList in os.walk(directory):
		for fName in fileList:
			if fName == fileName + extensions[0]: #if file is found/existing; index 0 could be anything
				fileExistInfo["exist"] = True 
			elif fName == fileName + extensions[1]: #index 1 will check for .TRB.xlsx extension
				fileExistInfo["exist"] = True 
				fileExistInfo["xlsx"] = True 
	return fileExistInfo
#this function converts .xls files to .xlsx for openpyxl to use
def convertExcel(filePath):
	try:
		excel = win32.gencache.EnsureDispatch('Excel.Application')
		wb = excel.Workbooks.Open(filePath)

		wb.SaveAs(filePath+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
		wb.Close()                               #FileFormat = 56 is for .xls extension
		excel.Application.Quit()

		del excel
		#delete old excel
		os.remove(filePath)
	except:
		os.system("TASKKILL /IM excel.exe")
		convertExcel(filePath)
#this function removes the corrupt excel file and saves it as a new one; returns the name with a space 
def removeCorruptExcel(filePath, fileName):
	newFileName = ""
	try:
		excel = win32.gencache.EnsureDispatch('Excel.Application')
		excel.DisplayAlerts = False #prevent from asking if you want to overwrite the JV if it already exists
		wb = excel.Workbooks.Open(filePath + "\\" + fileName, CorruptLoad = 1)

		for i in fileName:
			if i == "J":
				newFileName += " J"
			else:
				newFileName += i

		wb.SaveAs(filePath + "\\" + newFileName, FileFormat = 51)
		excel.DisplayAlerts = True #reset the display alerts
		wb.Close()
		excel.Application.Quit()

		del excel
	    #delete old excel
		os.remove(filePath + "\\" + fileName)
	except:
		os.system("TASKKILL /IM excel.exe")
		removeCorruptExcel(filePath, fileName)

	return newFileName
#this function will detect a change in the columns for certain headers like Debit, Credit, etc.
#it will also detect a change in the amount of VISA cards there are 
def checkTrbChange(curTrb, errorMsg):
	cardCount = 0
	isChanged = False
	#Used for making sure the cell is a header
	colNames = {"Debit": 0, "Credit": 0, "MTD Debit": 0, "MTD Credit": 0, "Count": 0, "MTD Count": 0}
	#2 arrays for the column letters in the CSV
	colLetters0 = []
	colLetters1 = []
	#2 arrays for the column letters in the new downloaded TRB
	colCoords0 = []
	colCoords1 = []

	#step 1. open csv for previously stored column letters
	with open('columns.csv') as csv_file:
		lineCount = 0
		csv_reader = csv.reader(csv_file, delimiter=',')
		for row in csv_reader:
			for col in range(6):
				if lineCount == 0:
					colLetters0.append(row[col])
				else:
					colLetters1.append(row[col])
			lineCount += 1
			
	#step 2. find the header cells and store the letters
	for i in range(1, curTrb.max_row + 1):
		for j in range(1, curTrb.max_column+1):
			curTrbVal = curTrb.cell(row=i, column=j).value
			#first char of coordinate ex: A in A4
			curTrbCoord = curTrb.cell(row=i, column=j).coordinate[0]
			#if its Debit, Credit, etc and if its not in one of the headers arrays already
			curTrbVal = str(curTrbVal)
			if curTrbVal in colNames:
				#increase counter for the amount of times the header occurs
				colNames[curTrbVal] += 1
				if curTrbCoord not in colCoords0 and colNames[curTrbVal] == 1:
					colCoords0.append(curTrbCoord)
				elif curTrbCoord not in colCoords1 and colNames[curTrbVal] == 3 and curTrbVal != "Count":
					colCoords1.append(curTrbCoord)
				elif curTrbCoord not in colCoords1 and colNames[curTrbVal] == 2 and curTrbVal == "MTD Count":
					colCoords1.append(curTrbCoord)
				elif curTrbCoord not in colCoords1 and colNames[curTrbVal] == 4 and curTrbVal == "Count":
					colCoords1.append(curTrbCoord)
			elif "VISA" in curTrbVal:
				cardCount += 1
	#step 3. compare column or letter positions
	for k in range(6):
		#if headers are not the same in sequence and if the values aren't equal
		if colCoords0[k] != colLetters0[k] or colCoords1[k] != colLetters1[k]:
			isChanged = True

	#update error message and columns if there is a change
	errorMsg.seek(0)
	errorMsg.truncate()
	errorLogMsgs = []
	if isChanged:
		message = "The columns have shifted. Template JV.xlsx's JV Sheet formulae must be updated"
		errorMsg.write(message)
		errorLogMsgs.append(message)
		with open('columns.csv', 'w', newline = '') as csv_file:
			csv_writer = csv.writer(csv_file, delimiter = ',')
			csv_writer.writerow(colCoords0)
			csv_writer.writerow(colCoords1)


	#if there are a different number of visa cards someone has to change the template
	with open('card amounts.txt') as txt_file:
		message = 'The amount of cards has changed. Template JV.xlsx must be updated with the right amount of cards.'
		if txt_file.readline() != str(cardCount):
			isChanged = True
			errorMsg.write(message)
			errorLogMsgs.append(message)

	#update log file if there is anything to be updated
	if len(errorLogMsgs) > 0:
		updateLogFile(errorLogMsgs)

	#reset reader to line 1 of error message.txt
	errorMsg.seek(0)

	return isChanged

#this function copies an entire sheet to another sheet
def copySheet(src_sheet, dest_sheet):
	for i in range(1, src_sheet.max_row+1):
		for j in range(1, src_sheet.max_column+1):
			dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

#this function exports a sheet to pdf 
def exportPdf(wb_path, path_to_pdf):
	try:
		excel = win32.gencache.EnsureDispatch("Excel.Application")
		excel.Visible = False

		wb = excel.Workbooks.Open(wb_path)

		#Select first and only sheet
		wb.Worksheets(1).Select()
		wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
		excel.Application.Quit()
		del excel
	except:
		os.system("TASKKILL /IM excel.exe")
		exportPdf(wb_path, path_to_pdf)

def exportCsv(csvPath, excelSheet):
	wb = pyxl.load_workbook(excelSheet, data_only = True) #re-open JV excel sheet in data only mode
	csv_sheet = wb["for CSV"]
	with open(csvPath, 'w', newline = '') as csv_file:
		csv_writer = csv.writer(csv_file, delimiter = ',')
		for i in range(1, csv_sheet.max_row + 1):
			row = []
			for j in range(1, csv_sheet.max_column+1):
				cell = csv_sheet.cell(row=i, column=j).value
				row.append(cell)
			if not 0 in row and not None in row:
				csv_writer.writerow(row)
			del row #remove from memory
	wb.close()

def main():

	#check for errors, if no error run the script
	file_error = open('error.txt', "r+")
	file_error_message = open('error message.txt', 'r+')

	#main function for script
	def createJV(date, isXlsx, trbPath, csvPath):
		#directory name
		TRB_EXPORT_PATH = trbPath
		CSV_EXPORT_PATH = csvPath 

		#file name
		newJv_path = date[-6:] + "JV.xlsx"
		RAW_DOWNLOADED_TRB_NAME = date + ".TRB.xls"
		DOWNLOADED_TRB_NAME = RAW_DOWNLOADED_TRB_NAME + "x"

		if not isXlsx: #if the TRB selected isnt .xlsx convert it to .xlsx
			convertExcel(TRB_EXPORT_PATH + "\\" + RAW_DOWNLOADED_TRB_NAME)

		file_curTrb = pyxl.load_workbook(TRB_EXPORT_PATH + "\\" + DOWNLOADED_TRB_NAME)
		sheet_curTrb = file_curTrb["2166"]

		#check if the format for the trb changed 
		if checkTrbChange(sheet_curTrb, file_error_message):
			writeErrorFile(file_error,'true')
		file_curTrb.close()
		del file_curTrb
		del sheet_curTrb

		#check if the user fixed the template, if they didn't it will be true
		file_error.seek(0)
		if file_error.read() == "false":
			#Load downloaded TRB
			downloadedTrb_wb = pyxl.load_workbook(TRB_EXPORT_PATH + "\\" + DOWNLOADED_TRB_NAME)
			downloaded_trb = downloadedTrb_wb["2166"] #load the sheet for the downloaded TRB
			
			#Create new workbook and the new TRB sheet
			newtrb_wb = pyxl.Workbook() #create new JV wb
			newtrb_sheet = newtrb_wb['Sheet']
			newtrb_sheet.title = "TRB"

			#copy downloaded TRB to sheet of new JV
			copySheet(downloaded_trb, newtrb_sheet)
			downloadedTrb_wb.close() #close downloaded TRB
			del downloadedTrb_wb
			del downloaded_trb

			#Load the template and sheets from template
			ttrb_wb = pyxl.load_workbook('Template JV.xlsx')
			jv_sheet = ttrb_wb["JV"]
			csv_sheet = ttrb_wb["for CSV"]

			#Create the JV and CSV sheets for the new JV
			newjv_sheet = newtrb_wb.create_sheet("JV")
			newcsv_sheet = newtrb_wb.create_sheet("for CSV")
			copySheet(jv_sheet, newjv_sheet)
			copySheet(csv_sheet, newcsv_sheet)
			ttrb_wb.close() #close template TRB
			del ttrb_wb
			del jv_sheet
			del csv_sheet

			#print to PDF
			exportPdf(TRB_EXPORT_PATH + "\\" + DOWNLOADED_TRB_NAME, TRB_EXPORT_PATH + "\\PDF\\" + date + " TRB.pdf")

			#save and close new JV 
			newtrb_wb.save(TRB_EXPORT_PATH + "\\" + newJv_path)
			newtrb_wb.close()
			del newtrb_wb
			del newtrb_sheet
			del newjv_sheet
			del newcsv_sheet
			newJv_path = removeCorruptExcel(TRB_EXPORT_PATH, newJv_path) #reopen excel to remove corrupt, change the name to have a space before " JV"

			#export CSV and reopen the JV that was created in data only mode
			exportCsv(CSV_EXPORT_PATH + "\\VantivTRBDaily.csv", TRB_EXPORT_PATH + "\\" + newJv_path)

			#return messages of success
			successMsgs = []
			successMsgs.append(f'SUCCESS: Created {newJv_path} in {TRB_EXPORT_PATH}\\{newJv_path}')
			successMsgs.append(f'SUCCESS: Overwrote VantivTRBDaily.csv in {CSV_EXPORT_PATH}\\VantivTRBDaily.csv' )
			print(successMsgs[0])
			print(successMsgs[1])
			#update log
			updateLogFile(successMsgs)

			return successMsgs

	#GUI programming starts here
	root  = Tk()
	root.title("Worldwide Pay VISA Automation Script")
	root.geometry("700x300")

	#variables
	inputDate = StringVar()
	inputPath = StringVar()
	errorMessage = StringVar()
	trbDirectory = StringVar()
	csvDirectory = StringVar()
	widgets = []

	#initialize error label and directory paths
	errorMessage.set(file_error_message.read())
	with open('trb export path.txt', 'r') as file_export_path:
		trbDirectory.set(file_export_path.read())
	with open('csv export path.txt', 'r') as file_export_path:
		csvDirectory.set(file_export_path.read())

	#this button/function updates the error file to be able to run the script
	def resolveError(errMsg, errLabel):
		writeErrorFile(file_error, 'false')
		file_error_message.seek(0)
		file_error_message.truncate()
		file_error_message.write("")
		errMsg.set("")
		errLabel.destroy()
		del errLabel
	#this button/function runs the actual script by calling the main function
	def start(date, errMsg, trbDirectory, csvDirectory):

		#destroy any previous labels (success label and does not exist label if it exists in the array)
		widgetsLen = len(widgets)
		if widgetsLen > 0:
			x = 0
			while x < widgetsLen:
				widgets[0].destroy()
				del widgets[0]
				x += 1
		
		#try get excel, if there's an error that means excel is not running
		try:
			excel = win32.GetActiveObject("Excel.Application")
			ctypes.windll.user32.MessageBoxW(0, "Please close excel before running", "Close running processes first", 0)
			return 0
		except:
			#check if the files in the directories given exist
			trbExistInfo = checkFileExist(date.get(), trbDirectory.get(), [".TRB.xls", ".TRB.xlsx"])
			csvExistInfo = checkFileExist("VantivTRBDaily", csvDirectory.get(), ['.csv', ""])

			#check if the PDF directory exists in the TRB directory, if not make it
			#also check if the TRB exists to make sure the PDF directory is in the right folder
			if not os.path.exists(trbDirectory.get() + "\\PDF") and trbExistInfo.get("exist"):
				os.makedirs(trbDirectory.get() + "\\PDF")

			if trbExistInfo.get("exist") and csvExistInfo.get("exist"):
				#create message after creating JV
				running_label = Label(root, text = "Running... Please wait.") #Create label to tell them its running
				running_label.place(x = 50, y = 220)
				running_label.config(font = ("Helvetica", 12))
				root.update()
				msgs = createJV(date.get(), trbExistInfo.get("xlsx"), trbDirectory.get(), csvDirectory.get())
				running_label.destroy() #destroy label after running 
				del running_label
				root.update()
				#show user that the file was made
				try: #if the columns shifted then the createJV function wont return an array, which will throw an error
					success_label0 = Label(root, text = msgs[0], fg = "green")
					success_label0.place(x = 50, y = 220)
					success_label0.config(font = ("Helvetica", 10))
					widgets.append(success_label0)
					success_label1 = Label(root, text = msgs[1], fg = "green")
					success_label1.place(x = 50, y = 250)
					success_label1.config(font = ("Helvetica", 10))
					widgets.append(success_label1)
				except:
					error_label0 = Label(root, text = "You did not change the JV Template. Once changed, click \"I have changed the template\"", fg = "red")
					error_label0.place(x = 50, y = 220)
					error_label0.config(font = ("Helvetica", 10))
					widgets.append(error_label0)
				#update the error message and update the tkinter interface
				file_error_message.seek(0)
				errMsg.set(file_error_message.read())
				root.update()
			else:
				if not trbExistInfo.get("exist"): #if the TRB doesn't exist
					dne_label0 = Label(root, text = "Error: File " + date.get() + ".TRB.xls or File " + date.get() + f".TRB.xlsx does not exist in {trbDirectory.get()}", fg = "red")
					dne_label0.place(x = 50, y = 220)
					dne_label0.config(font = ("Helvetica", 10))
					widgets.append(dne_label0)
				if not csvExistInfo.get("exist"): #if the csv doesn't exist in the path provided
					dne_label1 = Label(root, text = f"Error: File VantivTRBDaily.csv does not exist in {csvDirectory.get()}", fg = "red")
					dne_label1.place(x = 50, y = 250)
					dne_label1.config(font = ("Helvetica", 10))
					widgets.append(dne_label1)
	#this menu function opens a directory chooser to set the import/export path
	def setPath(setDirectory):
		root.directory = filedialog.askdirectory()

		#convert path to windows path with back slashes 
		path = ""
		for char in root.directory:
			if char == '/':
				path += "\\"
			else:
				path += char
		#update the export path
		if path != "":
			setDirectory.set(path)

	#WIDGETS
	#menu bar
	menubar = Menu(root)
	menubar.add_command(label="Set TRB Folder Path", command= lambda:setPath(trbDirectory))
	menubar.add_command(label="Set CSV Export Path", command= lambda:setPath(csvDirectory))
	#error message
	label = Message(root, textvariable = errorMessage, relief = RAISED, width = 400, fg = "red")
	label.config(font = ("Helvetica", 14))
	label.place(x = 150, y = 0)
	#buttons
	button = Button(root, text = "I have changed the template", command = lambda:resolveError(errorMessage, widgets[0])) #resolve error button
	button2 = Button(root, text = "Run Script", command = lambda:start(inputDate, errorMessage, trbDirectory, csvDirectory)) #start script button
	button.config(font = ("Helvetica", 10))
	button2.config(font = ("Helvetica", 10))
	button.place(x = 435, y = 150, width = 175, height = 40)
	button2.place(x = 295, y = 150, width = 100, height = 40)
	#Labels
	label2 = Message(root, text = "Enter the date:", width = 150) #date label
	label2.config(font = ("Helvetica", 16))
	label2.place(x = 100, y = 90)
	#text field
	textfield = Entry(root, textvariable = inputDate, width = 15) #text box for date
	textfield.config(font = ("Helvetica", 16))
	textfield.place(x = 275, y = 90)
	#root
	root.config(menu=menubar)
	root.mainloop()

	#update export paths before closing
	with open('trb export path.txt', 'w') as file_export_path:
		file_export_path.write(trbDirectory.get())

	with open('csv export path.txt', 'w') as file_export_path:
		file_export_path.write(csvDirectory.get())

	#close error txt files
	file_error.close()
	file_error_message.close()

if __name__ == '__main__':
	main()